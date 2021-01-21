#New file on maincontroller Version 1.2
#Dec 15 2020



from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side

import os
from os import listdir
from os.path import isfile, join


#Nhap cac thu vien de xu ly file xls,xlsx
import numpy as np
import pandas as pd
import xlrd
import xlwt
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment,PatternFill
import xlsxwriter
import pyexcel
import pyexcel_io





#Cac thu vien den copy file
import datetime
import shutil





#Xu ly file cua BO

#Doc danh sach file
input_path = 'input'
input_files = [f for f in listdir(input_path) if isfile(join(input_path,f))]

all_input_files = []
for item in input_files:
    str_input = input_path + '\\' + item
    all_input_files.append(str_input)


#Kiem tra lai danh sach file nhap vao
for item in all_input_files:
    print(item)




#Tien hanh lay du lieu tu file tong quat de xu ly cho bo
for tenfile in all_input_files:
    if tenfile.find("Tonghop") != -1: 
        book = xlrd.open_workbook(tenfile)
        for name in book.sheet_names():
            if name.find("Kết quả") != -1:
                #Kiem tra ten
                #print(book.sheet_by_name(name))
                
                #Tien hanh doc du lieu
                df = pd.read_excel(tenfile, sheet_name = name)
                
                #Check tinh trang sheet 
                #print("{0} {1} {2}".format(sh.name, sh.nrows, sh.ncols))
                #print(type(df))

                #Xu ly du lieu
                row = 4
                kq = True
                while kq:
                    try:
                        int(df.iloc[row,0])
                        row += 1    
                    except:
                        kq = False
                #Chot row & col de xu ly
                final_row = row - 1
                final_col = 20
                first_row = 2
                first_col = 0

                #checkinput
                #print(row)
                
                #check dữ liệu
                #col = 0
                #while col<50:
                #    try:
                #        print(df.iloc[2,col])
                #        col += 1
                #    except:
                #        print(col)
                #        break
                
                #Doc toan bo dataframe vao xu ly
                #Lay ten lop
                cut_tenfile = tenfile.split('_')
                lop = cut_tenfile[len(cut_tenfile)-1]
                lop = lop[:-4]
                #print(lop)

                #Tao file xuat
                now = str(datetime.datetime.now())[:19]
                now = now.replace(":","_")

                dst_dir="output\\output_bo\\bo_kqht_"+ lop + "_" + str(now) + ".xlsx"

                #Xoa truoc 2 dong dau cua dataframe
                df = df.iloc[2:]

                writer = pd.ExcelWriter(dst_dir,engine = "openpyxl", mode = 'w')
                df.to_excel(writer,sheet_name = 'Sheet1', startrow=2, startcol=0, header=False, index=False)
         
                
                writer.save()
                
                #Xu ly tieu de bang openpyxl
                wb = openpyxl.load_workbook(dst_dir)
                ws = wb.active

                font_title = openpyxl.styles.Font(name='Times New Roman',size = 11,bold=True,color= '00FF0000')
                font_marktitle = openpyxl.styles.Font(name='Times New Roman',size = 11,bold=True)
                style_title = openpyxl.styles.Alignment(horizontal='center', vertical='center',wrap_text= True)
                
                grey = "DDDDDD"
                black= "00808080"
                

                

                def headerDealer(col,row,strmerge,strip,fontchoice,backgroundchoice):
                    cell_dealing = ws.cell(column=col,row=row)
                    cell_dealing.value = strip
                    cell_dealing.font = fontchoice
                    ws.merge_cells(strmerge)
                    cell_dealing.alignment = style_title
                    cell_dealing.fill =  PatternFill("solid", fgColor=backgroundchoice)

                headerDealer(1,1,"A1:A2","STT",font_title,grey)
                headerDealer(2,1,"B1:B2","Mã lớp",font_title,grey)
                headerDealer(3,1,"C1:C2","Mã học sinh",font_title,grey)
                headerDealer(4,1,"D1:D2","Họ tên",font_title,grey)
                headerDealer(5,1,"E1:E2","Ngày sinh",font_title,grey)
                headerDealer(6,1,"F1:F2","Toán",font_marktitle,black)
                headerDealer(7,1,"G1:G2","Vật lý",font_marktitle,black)
                headerDealer(8,1,"H1:H2","Hóa học",font_marktitle,black)
                headerDealer(9,1,"I1:I2","Sinh học",font_marktitle,black)
                headerDealer(10,1,"J1:J2","Tin học",font_marktitle,black)
                headerDealer(11,1,"K1:K2","Ngữ văn",font_marktitle,black)
                headerDealer(12,1,"L1:L2","Lịch sử",font_marktitle,black)
                headerDealer(13,1,"M1:M2","Địa lý",font_marktitle,black)
                headerDealer(14,1,"N1:N2","Ngoại ngữ",font_marktitle,black)
                headerDealer(15,1,"O1:O2","Công nghệ",font_marktitle,black)
                headerDealer(16,1,"P1:P2","GD QP-AN",font_marktitle,black)
                headerDealer(17,1,"Q1:Q2","Thể dục",font_marktitle,black)
                headerDealer(18,1,"R1:S1","Tự chọn",font_marktitle,black)
                headerDealer(20,1,"T1:T2","GDCD",font_marktitle,black)
                headerDealer(21,1,"U1:U2","ĐTB các môn",font_marktitle,black)
                headerDealer(22,1,"V1:X1","Kết quả xếp loại và DH thi đua",font_marktitle,black)


                def withoutmerge_headerDealer(col,row,strip,fontchoice,backgroundchoice):
                    cell_dealing = ws.cell(column=col,row=row)
                    cell_dealing.value = strip
                    cell_dealing.font = fontchoice
                    cell_dealing.alignment = style_title
                    cell_dealing.fill =  PatternFill("solid", fgColor=backgroundchoice)

                withoutmerge_headerDealer(18,2,"Ngoại ngữ 2",font_marktitle,black)
                withoutmerge_headerDealer(19,2,"Nghề phổ thông",font_marktitle,black)
                withoutmerge_headerDealer(22,2,"Học lực",font_marktitle,black)
                withoutmerge_headerDealer(23,2,"Hạnh kiểm",font_marktitle,black)
                withoutmerge_headerDealer(24,2,"Danh hiệu thi đua",font_marktitle,black)


                


                #Def dich chuyen du lieu
                def move_range(startcol,startrow,endcol,endrow,rowmove,colmove):
                    str_range = startcol+str(startrow)+':'+endcol+str(endrow)
                    ws.move_range(str_range,rows=rowmove,cols=colmove)

                #Dich chuyen phan du lieu dataframe
                if lop.find('10') != -1:
                    move_range('C',3,'S',final_row+1,0,3)
                    move_range('B',3,'B',final_row+1,0,2)
                    move_range('T',3,'V',final_row+1,0,2)
                    move_range('S',3,'S',final_row+1,0,2)
                    move_range('O',3,'O',final_row+1,0,5)
                    move_range('P',3,'P',final_row+1,0,-1)
                    move_range('R',3,'R',final_row+1,0,-2)
                if lop.find('12') != -1:
                    move_range('C',3,'S',final_row+1,0,3)
                    move_range('B',3,'B',final_row+1,0,2)
                    move_range('T',3,'V',final_row+1,0,2)
                    move_range('S',3,'S',final_row+1,0,2)
                    move_range('O',3,'O',final_row+1,0,5)
                    move_range('P',3,'P',final_row+1,0,-1)
                    move_range('R',3,'R',final_row+1,0,-2)
                if lop.find('11') != -1:
                    move_range('C',3,'T',final_row+1,0,3)
                    move_range('B',3,'B',final_row+1,0,2)
                    move_range('U',3,'W',final_row+1,0,1)
                    move_range('T',3,'T',final_row+1,0,1)
                    move_range('O',3,'O',final_row+1,0,5)
                    move_range('P',3,'P',final_row+1,0,-1)
                    move_range('R',3,'R',final_row+1,0,-2)



                #Dat lai cot dtb lay 1 so sau dau phay
                for row in range(1, final_row+2):
                    ws["F{}".format(row)].number_format = '#,#0.0'
                    ws["G{}".format(row)].number_format = '#,#0.0'
                    ws["H{}".format(row)].number_format = '#,#0.0'
                    ws["I{}".format(row)].number_format = '#,#0.0'
                    ws["J{}".format(row)].number_format = '#,#0.0'
                    ws["K{}".format(row)].number_format = '#,#0.0'
                    ws["L{}".format(row)].number_format = '#,#0.0'
                    ws["M{}".format(row)].number_format = '#,#0.0'
                    ws["N{}".format(row)].number_format = '#,#0.0'
                    ws["O{}".format(row)].number_format = '#,#0.0'
                    ws["P{}".format(row)].number_format = '#,#0.0'
                    ws["R{}".format(row)].number_format = '#,#0.0'
                    ws["S{}".format(row)].number_format = '#,#0.0'
                    ws["T{}".format(row)].number_format = '#,#0.0'
                    ws["U{}".format(row)].number_format = '#,#0.0'
                    


                #Dong khung 
                def __format_ws__(ws, cell_range):
                    border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

                    rows = ws[cell_range]
                    for row in rows:
                        for cell in row:
                            cell.border = border

                cells_range = 'A1:X'+str(final_row+1)
                __format_ws__(ws,cells_range)

                #Di kiem tra va lay du lieu cac cot ma lop, ma hoc sinh, ngay thang nam sinh
                for fls in all_input_files:
                    if fls.find('KQHT') != -1:
                        #Doc file 
                        idbook = pd.read_excel(fls,sheet_name = 'Sheet1')
                        
                        #Kiem tra lop co thoa man khong?
                        #print(idbook.loc[1][1])
                        #print(lop)
                        #kq = True
                        #if lop != idbook.loc[1][1]:
                        #    kq = False
                        #print(kq)

                        if idbook.loc[1][1] == lop: 
                            #Xuly lay dataframe nay de duyet
                            idmf = idbook

                #Da lay duoc du lieu dataframe vao dataframe ten idmf
                #Chuyen dataframe nay sang list de de xuly voi openpyxl
                listid = idmf.values.tolist()
                listid = listid[1:]
                #print(listid)

                #Tao thu tuc kiem tra va push du lieu len file excels
                def write_cell_data(input_row,input_col,input_data):
                    cell_dealing = ws.cell(column=input_col,row=input_row)
                    cell_dealing.value = input_data

                #Load du lieu len file
                for rec_index in range(0,len(listid)):
                    #print(listid[rec_index])
                    write_cell_data(rec_index+3,1,listid[rec_index][0])
                    write_cell_data(rec_index+3,2,listid[rec_index][1])
                    write_cell_data(rec_index+3,3,listid[rec_index][2])
                    write_cell_data(rec_index+3,5,listid[rec_index][4])

                wb.save(dst_dir)
#Da hoan tat xu ly file cua bo giao duc

                
                
#Tien hanh xu ly file xuat cua so giao duc
#Chay file de truy xuat du lieu
for tenfile in all_input_files:
    #print(tenfile)
    
    #Tien hanh nhap file so_nhapdiemchitiet lay du lieu ID HS cua lop 
    #(GV phai download toan bo cac file nhap diem cua tung lop nay cho vao ds id)
    if (tenfile.find('Nhap diem chi tiet mon hoc') != -1):
        #Load workbook & Check title
        so_bk = xlrd.open_workbook(tenfile)
        df_lop = pd.read_excel(so_bk,sheet_name= 'Sheet1')
        
        #Iterate Title & Take Class name
        #print(df_lop.iat[2,0])
        raw_class_name = df_lop.iat[2,0]
        class_name_sp = raw_class_name.split('-')
        class_name = class_name_sp[1][1:-1]
        
        #Test Permited class_name???
        #print('*'+class_name+'*')
        #Da hoan tat lay class name 

        #Xu ly thu tuc lay dataframe chuan bi tao file moi
        df_lop = df_lop[6:]
        df_lop.drop(df_lop.columns[[6,7,8,9,10,11,12]],inplace = True, axis = 1)

        #print(df_lop)

        #Tao list mon
        tenmon_infiles = ['toan_hoc','tin_hoc','vat_li','hoa_hoc','sinh_hoc','lich_su','dia_li','ngu_van','ngoai_ngu','gdcd','cong_nghe']
        tenmon_view = ['TOÁN','TIN HỌC','VẬT LÍ','HÓA HỌC','SINH HỌC','LỊCH SỬ','ĐỊA LÍ','NGỮ VĂN','NGOẠI NGỮ','GIÁO DỤC CÔNG DÂN','CÔNG NGHỆ']

        for mon in range(0,len(tenmon_infiles)):
            #Tao file name
            so_op_name = 'so_nhapdiemchitiet_' + class_name + '_mon_' + tenmon_infiles[mon]

            #Cat thoi diem
            now = str(datetime.datetime.now())[:19]
            now = now.replace(':','_')
            now = now.replace('-','_')
            now = now.replace(' ','_')

            so_op_name = so_op_name + '_' + now + '.xlsx'
            #print(tenfile)

            #Chuan bi file moi va nhap lieu vao file moi
            dst_so_dir = "output\\output_so\\" + so_op_name
            #print(dst_so_dir)

            writer_so = pd.ExcelWriter(dst_so_dir,engine = "openpyxl", mode = 'w')
            df_lop.to_excel(writer_so,sheet_name = 'Sheet1', startrow=7, startcol=0, header=False, index=False)

            writer_so.close()
            
            #Tien hanh mo lai file va xu ly
            wb_so = openpyxl.load_workbook(filename=dst_so_dir)
            ws_so = wb_so.active
            
            so_cell = ws_so.cell(column=1,row=1)
            so_cell.value = 'Sở giáo dục và đào tạo'
            truong_cell = ws_so.cell(column=1,row=2)
            truong_cell.value = 'Đơn vị: THPT Đức Hòa'

            #Hoan tat load Tieu de truong & so 

            #Load du lieu diem
            khoilop = class_name[:2]
            #print(khoilop)
            for mark_file_search in all_input_files:
                if ((mark_file_search.find(tenmon_infiles[mon]) != -1) and (mark_file_search.find(str(khoilop)) != -1)):
                    #Doc file chua diem mon cua lop 
                    sheet_takepoint = tenmon_infiles[mon]+'_'+class_name.lower()
                    df_pts_class = pd.read_excel(mark_file_search,sheet_name=sheet_takepoint)
                    #print(df_pts_class)
                    #print(sheet_takepoint)
                    pts_of_class = df_pts_class.values.tolist()
                    pts_of_class = pts_of_class[4:-7]
                    #print(pts_of_class)
                    #Hoan thanh load du lieu xuong


            for row in range(7, len(pts_of_class)+7):
                ws_so["G{}".format(row)].number_format = '#,#0.0'
                ws_so["H{}".format(row)].number_format = '#,#0.0'
                ws_so["I{}".format(row)].number_format = '#,#0.0'
                ws_so["J{}".format(row)].number_format = '#,#0.0'
                ws_so["K{}".format(row)].number_format = '#,#0.0'
                ws_so["L{}".format(row)].number_format = '#,#0.0'
                ws_so["M{}".format(row)].number_format = '#,#0.0'
                    

            #Load du lieu len file
            def write_pts_data(iprow,ipcol,ipdata):
                ws_so.cell(row=iprow, column=ipcol).value = ipdata
                

            for row_data in range(2,len(pts_of_class)):
                ws_so.cell(row=row_data+6,column=5).number_format = 'dd/mm/yyyy'

                check_point_1 = False
                check_point_2 = False
                check_point_3 = False
                check_point_4 = False


                if pts_of_class[1][4] == 'TX1':
                    write_pts_data(row_data+6,7,pts_of_class[row_data][4])
                    check_point_1 = True
                    #print(pts_of_class[row_data][4])

                if pts_of_class[1][5] == 'TX2':
                    write_pts_data(row_data+6,8,pts_of_class[row_data][5])
                    check_point_2 = True
                    #print(pts_of_class[row_data][5])

                if pts_of_class[1][6] == 'TX3':
                    write_pts_data(row_data+6,9,pts_of_class[row_data][6])
                    check_point_3 = True
                    #print(pts_of_class[1][6])
                    #print(pts_of_class[row_data][6])

                if pts_of_class[1][7] == 'TX4':
                    write_pts_data(row_data+6,10,pts_of_class[row_data][7])
                    check_point_4 = True
                    #print(pts_of_class[1][7])
                    #print(pts_of_class[row_data][7])

                print(pts_of_class)
                for item in range(0,len(pts_of_class[0])):
                    if pts_of_class[0][item] == 'ĐĐGgk':
                        index_gk = item
                #index_gk = pts_of_class[0].index('gk')
                #print(index_gk)
                input_col = 7
                if check_point_1:
                    input_col += 1 
                if check_point_2:
                    input_col += 1 
                if check_point_3:
                    input_col += 1 
                if check_point_4:
                    input_col += 1 


                write_pts_data(row_data+6,11,pts_of_class[row_data][index_gk])
                write_pts_data(row_data+6,12,pts_of_class[row_data][index_gk+1])
                write_pts_data(row_data+6,13,pts_of_class[row_data][index_gk+2])



                #print(str(row_data)+ '_5_' + str(pts_of_class[row_data][5]))
                #print(str(row_data)+ '_6_' + str(pts_of_class[row_data][6]))
                #print(str(row_data)+ '_7_' + str(pts_of_class[row_data][7]))
                
                #print('*'+pts_of_class[1][4]+'*')
                #print('*'+pts_of_class[1][5]+'*')
                #write_pts_data(row_data+6,7,pts_of_class[row_data][5])




            #Build he thong tieu de trang
            def headerDealer_nobackground(col,row,strmerge,strip,fontchoice):
                    cell_dealing = ws_so.cell(column=col,row=row)
                    cell_dealing.value = strip
                    cell_dealing.font = fontchoice
                    ws_so.merge_cells(strmerge)
                    cell_dealing.alignment = style_title

            alice_blue = "10edef"
            light_yellow = "ebfb76"
            def headerDealer_so(col,row,strmerge,strip,fontchoice,backgroundchoice):
                    cell_dealing = ws_so.cell(column=col,row=row)
                    cell_dealing.value = strip
                    cell_dealing.font = fontchoice
                    ws_so.merge_cells(strmerge)
                    cell_dealing.alignment = style_title
                    cell_dealing.fill =  PatternFill("solid", fgColor=backgroundchoice)


            def headerDealer_withoutmerge_so(col,row,strip,fontchoice,backgroundchoice):
                    cell_dealing = ws_so.cell(column=col,row=row)
                    cell_dealing.value = strip
                    cell_dealing.font = fontchoice
                    cell_dealing.alignment = style_title
                    cell_dealing.fill =  PatternFill("solid", fgColor=backgroundchoice)

            def headerDealer_withoutmerge_withoutcolor_so(col,row,strip,fontchoice):
                    cell_dealing = ws_so.cell(column=col,row=row)
                    cell_dealing.value = strip
                    cell_dealing.font = fontchoice
                    cell_dealing.alignment = style_title
                    


            headerDealer_nobackground(1,6,'A6:A7','STTHS',font_marktitle)
            headerDealer_nobackground(2,6,'B6:B7','Mã học sinh',font_marktitle)
            headerDealer_nobackground(3,6,'C6:C7','Họ đệm',font_marktitle)
            headerDealer_nobackground(4,6,'D6:D7','Tên',font_marktitle)
            headerDealer_nobackground(5,6,'E6:E7','Ngày sinh',font_marktitle)
            headerDealer_nobackground(6,6,'F6:F7','Trạng thái học',font_marktitle)
            headerDealer_so(7,6,'G6:J6','Điểm kiểm tra thường xuyên',font_marktitle,alice_blue)
            headerDealer_withoutmerge_withoutcolor_so(7,7,'KTTX1',font_marktitle)
            headerDealer_withoutmerge_withoutcolor_so(8,7,'KTTX2',font_marktitle)
            headerDealer_withoutmerge_withoutcolor_so(9,7,'KTTX3',font_marktitle)
            headerDealer_withoutmerge_withoutcolor_so(10,7,'KTTX4',font_marktitle)
            headerDealer_withoutmerge_withoutcolor_so(11,7,'KTGK',font_marktitle)
            headerDealer_withoutmerge_so(11,6,'Điểm kiểm tra giữa kỳ',font_marktitle,light_yellow)
            headerDealer_so(12,6,'L6:L7','Điểm KT HK',font_marktitle,light_yellow)
            headerDealer_nobackground(13,6,'M6:M7','Điểm TB',font_marktitle)


            #Dong khung 
            def __format_ws_so__(ws, cell_range):
                border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))

                rows = ws_so[cell_range]
                for row in rows:
                    for cell in row:
                        cell.border = border

            cells_range = 'A6:M'+str(len(pts_of_class)+5)
            __format_ws_so__(ws_so,cells_range)

            #Tao tieu de file 
            title_file = 'NHẬP ĐIỂM CHI TIẾT MÔN HỌC - ' + class_name.upper() + '_' + tenmon_view[mon]
            headerDealer_nobackground(1,4,'A4:M4',title_file,font_marktitle)

            wb_so.save(dst_so_dir)

            #print(dst_so_dir)
            xls_dest = dst_so_dir.split('\\')
            xls_dest[1] = 'output_so_xls'
            xls_dest[2] = xls_dest[2][:-1]
            #print(xls_dest)

            new_dest_so = xls_dest[0] + '\\' + xls_dest[1] + '\\' + xls_dest[2]
            #print(new_dest_so)
            
            pyexcel.save_book_as(file_name = dst_so_dir,dest_file_name = new_dest_so)























                #atest = ws['A8']
                #print(atest.value)

                

#Mau xu ly
#book = xlrd.open_workbook("input\Sodiem_Tonghop_10A1.xls")
#print("The number of worksheets is {0}".format(book.nsheets))
#print("Worksheet name(s): {0}".format(book.sheet_names()))
#sh = book.sheet_by_index(0)
#print("{0} {1} {2}".format(sh.name, sh.nrows, sh.ncols))

#for rx in range(sh.nrows):
#    print(sh.row(rx))


#Tu lieu xu ly cho file cua so 

  #src_dir0="Not_touch_core\\so_nhapdiemchitiet.xls"
                #src_dir="Not_touch_core\\so_nhapdiemchitiet.xlsx"
                #dst_dir="output\\output_so\\so_nhapdiemchitiet_"+ lop + "_" + str(now) + ".xlsx"

#str_school_title = 'NHẬP ĐIỂM CHI TIẾT MÔN LỚP ' + lop
                
                #style_header = openpyxl.styles.Font(name='Times New Roman',sz=14,b=True)
                #style_center = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                #normalstyle = openpyxl.styles.Font(name='Times New Roman')

                #wb = openpyxl.load_workbook(dst_dir)
                #ws = wb.active

                #cell_so = ws.cell(row=1,column=1)
                #cell_so.value = 'Sở giáo dục và đào tạo'
                #cell_so.font = normalstyle
                #cell_truong = ws.cell(row=2,column=1)
                #cell_truong.value = 'Đơn vị: THPT Đức Hòa'
                #cell_truong.font = normalstyle
                

                ##Title lop
                #cell_title = ws.cell(row=4,column=1)
                #cell_title.value = str_school_title
                #cell_title.font = style_header
                #cell_title.alignment = style_center
                #ws.merge_cells("A4:M4")
                #wb.save(dst_dir)













#Raw procedures

#Xu ly tieu de
                #df_upper_title = pd.DataFrame({'Data':['Sở giáo dục và đào tạo','Đơn vị: THPT Đức Hòa']})
                #df_upper_title.to_excel(writer, startcol=0,startrow=0, header=None, index=False)
                #str_school_title = 'NHẬP ĐIỂM CHI TIẾT MÔN LỚP ' + lop

                #str_school_title = 'NHẬP ĐIỂM CHI TIẾT MÔN LỚP ' + lop
                #list_sc_title = list([str_school_title]*19)
                #print(list_sc_title)
                #df_class = pd.DataFrame([list_sc_title])
             
                #Dang bi vuong tieu de lop
                #Style option: 'color': font-weight: bold lua chon de xu ly bold face
                #df_class.style.set_properties(color="yellow")
                
                #Sap vi tri tieu de 
                #df_class.style.set_properties(**{'font-size':'14pt', 'text-align':'center','font-weight':'bold'}).to_excel(writer, startcol=0, startrow=3, header = None, index=False)
                
                #Luu du lieu lai



#Code cu xu ly di chuyen dong lenh
#st_range = 'C3:S'+ str(final_row+1)
                #print(st_range)
                #ws.move_range(st_range,rows=0,cols=3)

                #name_range = 'B3:B'+str(final_row+1)
                #ws.move_range(name_range,rows=0,cols=2)

                #kqxl_range = 'T3:V'+str(final_row+1)
                #ws.move_range(kqxl_range,rows=0,cols=2)

                #kqxl_range = 'S3:S'+str(final_row+1)
                #ws.move_range(kqxl_range,rows=0,cols=2)
